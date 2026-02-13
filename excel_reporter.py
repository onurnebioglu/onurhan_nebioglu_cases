import os
import subprocess
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExcelReporter:
    

    _BLUE = "1F4E79"
    _WHITE = "FFFFFF"
    _GREEN_BG = "C6EFCE"
    _GREEN_FG = "006100"
    _RED_BG = "FFC7CE"
    _RED_FG = "9C0006"
    _GREY_BG = "F2F2F2"
    _GREY_FG = "808080"
    _BORDER = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )

    def __init__(self, path="test_report.xlsx"):
        self.path = os.path.abspath(path)
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Test Report"
        self._data_start = 5
        self._step_count = 0

    

    def add_steps(self, steps: list):
        """Register all steps as Pending and write the initial spreadsheet."""
        ws = self.ws
        widths = {"A": 6, "B": 55, "C": 14, "D": 65, "E": 12}

     
        ws.merge_cells("A1:E1")
        title = ws["A1"]
        title.value = "INSIDER QA ENGINEER ASSESSMENT"
        title.font = Font(bold=True, size=14, color=self._WHITE)
        title.fill = PatternFill("solid", fgColor=self._BLUE)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        
        ws.merge_cells("A2:E2")
        sub = ws["A2"]
        sub.value = f"Test Execution — {datetime.now().strftime('%d %B %Y, %H:%M:%S')}"
        sub.font = Font(italic=True, size=10, color="444444")
        sub.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 20

        ws.row_dimensions[3].height = 6

      
        headers = ["#", "Test Step", "Status", "Details", "Time"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = Font(bold=True, color=self._WHITE, size=10)
            cell.fill = PatternFill("solid", fgColor=self._BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._BORDER
        ws.row_dimensions[4].height = 22

        for letter, w in widths.items():
            ws.column_dimensions[letter].width = w
        ws.freeze_panes = "A5"

       
        self._step_count = len(steps)
        for i, desc in enumerate(steps, 1):
            row = self._data_start + i - 1
            for col in range(1, 6):
                c = ws.cell(row=row, column=col)
                c.fill = PatternFill("solid", fgColor=self._GREY_BG)
                c.font = Font(color=self._GREY_FG, size=10)
                c.border = self._BORDER
            ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=desc)
            ws.cell(row=row, column=3, value="Pending").alignment = Alignment(horizontal="center")

        self._write_summary()
        self.wb.save(self.path)

    

    def pass_step(self, step: int, details: str = ""):
        self._set(step, "PASSED", details, self._GREEN_BG, self._GREEN_FG)

    def fail_step(self, step: int, details: str = ""):
        self._set(step, "FAILED", details, self._RED_BG, self._RED_FG)

    def _set(self, step, status, details, bg, fg):
        row = self._data_start + step - 1
        ws = self.ws
        ws.cell(row=row, column=3, value=status)
        ws.cell(row=row, column=4, value=str(details)[:250])
        ws.cell(row=row, column=5, value=datetime.now().strftime("%H:%M:%S"))
        for col in range(1, 6):
            c = ws.cell(row=row, column=col)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(color=fg, size=10, bold=(col == 3))
            c.border = self._BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        self._write_summary()
        self.wb.save(self.path)


    def _write_summary(self):
        ws = self.ws
        r = self._data_start + self._step_count + 1
        passed = sum(
            1 for i in range(self._step_count)
            if ws.cell(row=self._data_start + i, column=3).value == "PASSED"
        )
        failed = sum(
            1 for i in range(self._step_count)
            if ws.cell(row=self._data_start + i, column=3).value == "FAILED"
        )
        pending = self._step_count - passed - failed

        ws.merge_cells(f"A{r}:B{r}")
        summary = ws.cell(row=r, column=1)
        summary.value = "SUMMARY"
        summary.font = Font(bold=True, size=11, color=self._BLUE)
        summary.alignment = Alignment(horizontal="right")

        ws.cell(row=r, column=3, value=f"{passed} Passed").font = Font(
            color=self._GREEN_FG, bold=True, size=10
        )
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4).value = (
            f"{failed} Failed / {pending} Pending" if (failed or pending)
            else "All steps passed!"
        )
        ws.cell(row=r, column=4).font = Font(
            color=self._RED_FG if failed else self._GREEN_FG, bold=True, size=10
        )

    

    def open_file(self):
        """Open the final Excel report in the default spreadsheet app."""
        subprocess.Popen(["open", self.path])
